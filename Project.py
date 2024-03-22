import mysql.connector
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl import *


con = mysql.connector.connect(
    host="localhost", username="root", password="Narenguru2007", database="project"
)

mycur = con.cursor()
mycur.execute(
    "create table if not exists student(admn int primary key, sname varchar(50) not null, fname varchar(50) not null, mname varchar(50), sclass int not null, ssec varchar(5) not null)"
)
con.commit()
mycur.execute(
    "create table if not exists marks(name varchar(50), admn int primary key, sclass int not null, ssec varchar(5), sub1 float(5,2), sub2 float(5,2), sub3 float(5,2), sub4 float(5,2), sub5 float(5,2), total float(5,2), average float(5,2)"
)
con.commit()


def newstudent():
    cur = con.cursor()
    admn = int(input("Enter the Admission no of student:"))
    name = input("Enter the Name of the Student:").title()
    fname = input("Enter the Father name of the student:").title()
    mname = input("Enter the Mother name of the student:").title()
    sclass = int(input("Enter the Class of student (in number):"))
    sec = input("Enter the Section of the student:").upper()
    sql = f"insert into student values({admn},'{name}','{fname}','{mname}',{sclass},'{sec}')"
    cur.execute(sql)
    con.commit()


def mark():
    cur = con.cursor()
    name = input("Enter the Name of the Student:").title()
    admn = int(input("Enter the Admission no of student:"))
    sclass = int(input("Enter the Class of student (in number):"))
    sec = input("Enter the Section of the student:").upper()
    if sclass in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        sub1 = int(input("Enter the English mark:"))
        sub2 = int(input("Enter the Maths mark:"))
        sub3 = int(input("Enter the Tamil mark:"))
        sub4 = int(input("Enter the Science mark:"))
        sub5 = int(input("Enter the Social( or EVS) mark:"))
    elif sclass in [11, 12]:
        sub1 = int(input("Enter the English mark:"))
        sub2 = int(input("Enter the Maths/Psychology/CS mark:"))
        sub3 = int(input("Enter the Physics/B.St/Sociology mark:"))
        sub4 = int(input("Enter the Chemistry/Accountancy/History mark:"))
        sub5 = int(input("Enter the CS/Biology/Economics mark:"))
    else:
        print("Invaild Class")
    sql = f"insert into marks values('{name}',{admn},{sclass},'{sec}',{sub1},{sub2},{sub3},{sub4},{sub5},{sub1+sub2+sub3+sub4+sub5},{(sub1+sub2+sub3+sub4+sub5)/5}"
    cur.execute(sql)
    con.commit()


def getmarksclasssec():
    cur = con.cursor()
    sclass = int(input("Enter the class in number:"))
    sec = input("Enter the section to get excel:").upper()
    sql = f"select * from marks where sclass = {sclass} and ssec = '{sec}' "
    cur.execute(sql)
    res = tuple(cur.fetchall())
    head = ("Name", "Admn", "Class", "Sec", "sub1", "sub2", "sub3", "sub4", "sub5","total","Average")
    tup = (head,)
    tup = tup + res
    try:
        wb = load_workbook(filename=f"{sclass} {sec} marks.xlsx")
    except:
        wb = Workbook()
    sheet = wb.active
    i = 0
    for row in tup:
        i += 1
        j = 1
        for col in row:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1

    wb.save(filename=f"{sclass} {sec} marks.xlsx")
    print("File Saved.")


def getmarksclass():
    cur = con.cursor()
    sclass = int(input("Enter the class in number:"))
    sql = f"select * from marks where sclass = {sclass}"
    cur.execute(sql)
    res = tuple(cur.fetchall())
    head = ("Name", "Admn", "Class", "Sec", "sub1", "sub2", "sub3", "sub4", "sub5","total","average")
    tup = (head,)
    tup = tup + res
    try:
        wb = load_workbook(filename=f"{sclass} marks.xlsx")
    except:
        wb = Workbook()
    sheet = wb.active
    i = 0
    for row in tup:
        i += 1
        j = 1
        for col in row:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1

    wb.save(filename=f"{sclass} marks.xlsx")
    print("File Saved.")


def getbioclasssec():
    cur = con.cursor()
    sclass = int(input("Enter the class in number:"))
    sec = input("Enter the section to get excel:").upper()
    sql = f"select * from marks where sclass = {sclass} and ssec = '{sec}' "
    cur.execute(sql)
    res = tuple(cur.fetchall())
    head = ("Admn", "Name", "Father name", "Mother Name", "Class", "Sec")
    tup = (head,)
    tup = tup + res
    try:
        wb = load_workbook(filename=f"{sclass} {sec} marks.xlsx")
    except:
        wb = Workbook()
    sheet = wb.active
    i = 0
    for row in tup:
        i += 1
        j = 1
        for col in row:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1

    wb.save(filename=f"{sclass} {sec} marks.xlsx")
    print("File Saved.")


def getbioclass():

    cur = con.cursor()
    sclass = int(input("Enter the class in numbers:"))
    sql = f"select * from student where sclass = {sclass}"
    cur.execute(sql)
    res = tuple(cur.fetchall())
    head = ("Admn", "Name", "Father name", "Mother Name", "Class", "Sec")
    tup = (head,)
    tup = tup + res
    try:
        wb = load_workbook(filename=f"{sclass} biodata.xlsx")
    except:
        wb = Workbook()
    sheet = wb.active
    i = 0
    for row in tup:
        i += 1
        j = 1
        for col in row:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1

    wb.save(filename=f"{sclass} biodata.xlsx")
    print("File Saved.")


def getexcel():
    print("1.Enter 1 to get excel of students bio data classwise")
    print("2.Enter 2 to get excel of students bio data class and sectionwise")
    print("3.Enter 3 to get excel of student mark list classwise")
    print("4.Enter 4 to get excel of student mark list class and sectionwise")
    ch = int(input("Enter the choice:"))
    if ch == 1:
        getbioclass()
    elif ch == 2:
        getbioclasssec()
    elif ch == 3:
        getmarksclass()
    elif ch == 4:
        getmarksclasssec()
    else:
        print("Invalid choice")


def disstudmarkclass():
    cur = con.cursor()
    sclass = int(input("Enter the class in numbers:"))
    head = ["Name", "Admn", "Class", "Sec", "sub1", "sub2", "sub3", "sub4", "sub5"]
    sql = f" select * from marks where sclass = {sclass} "
    cur.execute(sql)
    res = cur.fetchall()
    print(tabulate(res, headers=head))


def disstudmarkclasssec():
    cur = con.cursor()
    sclass = int(input("Enter the class in numbers:"))
    sec = input("Enter the section:")
    head = ["Name", "Admn", "Class", "Sec", "sub1", "sub2", "sub3", "sub4", "sub5"]
    sql = f" select * from marks where sclass = {sclass} and ssec = '{sec}' "
    cur.execute(sql)
    res = cur.fetchall()
    print(tabulate(res, headers=head))


def disstudbioclass():
    cur = con.cursor()
    sclass = int(input("Enter the class in numbers:"))
    head = ["Admn", "Name", "Father name", "Mother Name", "Class", "Sec"]
    sql = f" select * from student where sclass={sclass} "
    cur.execute(sql)
    res = cur.fetchall()
    print(tabulate(res, headers=head))


def disstudbioclasssec():
    cur = con.cursor()
    sclass = int(input("Enter the class in numbers:"))
    sec = input("Enter the section:")
    head = ["Admn", "Name", "Father name", "Mother Name", "Class", "Sec"]
    sql = f" select * from student where sclass = {sclass} and ssec = '{sec}' "
    cur.execute(sql)
    res = cur.fetchall()
    print(tabulate(res, headers=head))


def display():
    print("1.Enter 1 to display students biodata")
    print("2.Enter 2 to display students marks")
    ch = int(input("Enter the choice in numbers:"))
    if ch == 1:
        print("1. Enter 1 to display students biodata classwise")
        print("2. Enter 2 to display students biodata class and section wise")
        ch1 = int(input("Enter the choice in numbers:"))
        if ch1 == 1:
            disstudbioclass()
        elif ch1 == 2:
            disstudbioclasssec()
        else:
            print("Invaild Choice")
    elif ch == 2:
        print("1. Enter 1 to display students marks classwise")
        print("2. Enter 2 to display students marks class and section wise")
        ch2 = int(input("Enter the choice in numbers:"))
        if ch2 == 1:
            disstudmarkclass()
        elif ch2 == 2:
            disstudmarkclasssec()
        else:
            print("Invails choice")
    else:
        print("Invaild choice")


while True:
    print("1. Enter 1 to add new student")
    print("2. Enter 2 for adding mark of a student")
    print("3. Enter 3 for Display details")
    print("4. Enter 4 for getting Excel details")
    print("5. Enter 5 to Exit")
    ch = int(input("Enter the choice in numbers:"))
    if ch == 1:
        newstudent()
    elif ch == 2:
        mark()
    elif ch == 3:
        display()
    elif ch == 4:
        getexcel()
    elif ch == 5:
        break
    else:
        print("Invalid Choice")
